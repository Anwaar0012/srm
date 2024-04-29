from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404, reverse
from django.http import HttpResponse
from django.views import View
from .models import LineItem, Invoice,Recovery
from .forms import LineItemFormset, InvoiceForm, RecoveryForm
from myapp.models import Product, Shop
from django.utils import timezone
# from django.contrib import messages
from django.contrib import messages
from datetime import datetime


import pdfkit

class InvoiceListView(View):
    def get(self, *args, **kwargs):
        invoices = Invoice.objects.all().order_by("-id")
        context = {
            "invoices":invoices,
        }

        return render(self.request, 'invoice/invoice-list.html', context)
    
    def post(self, request):        
        # import pdb;pdb.set_trace()
        invoice_ids = request.POST.getlist("invoice_id")
        invoice_ids = list(map(int, invoice_ids))

        update_status_for_invoices = int(request.POST['status'])
        invoices = Invoice.objects.filter(id__in=invoice_ids)
        # import pdb;pdb.set_trace()
        if update_status_for_invoices == 0:
            invoices.update(status=False)
        else:
            invoices.update(status=True)

        return redirect('invoice:invoice-list')

def createInvoice(request):
    """
    Invoice Generator page it will have Functionality to create new invoices, 
    this will be protected view, only admin has the authority to read and make
    changes here.
    """

    heading_message = 'Formset Demo'
    if request.method == 'GET':
        formset = LineItemFormset(request.GET or None)
        form = InvoiceForm(request.GET or None)
    elif request.method == 'POST':
        formset = LineItemFormset(request.POST)
        form = InvoiceForm(request.POST)
        
        if form.is_valid():
    
            
            invoice = Invoice.objects.create(customer=form.data["customer"],
                    customer_email=form.data["customer_email"],
                    billing_address = form.data["billing_address"],
                    date=form.data["date"],
                    due_date=form.data['due_date'],
                    message=form.data["message"],
                    salesperson=form.data["salesperson"],
                    manager=form.data["manager"],
                    sale_types=form.data['sale_types'],
                    routing=form.data["routing"],
                    paid_amount=form.data["paid_amount"],
                    # balance=form.data["balance"],
                    # previous_balance=form.data['previous_balance']
                    )
            # invoice.save()
            
        if formset.is_valid():
            # import pdb;pdb.set_trace()
            # extract name and other data from each form and save
            total = 0
            for form in formset:
                service = form.cleaned_data.get('service')
                # description = form.cleaned_data.get('description')
                quantity = form.cleaned_data.get('quantity')
                rate = form.cleaned_data.get('rate')
                if service and quantity and rate:
                    amount = Decimal(str(rate)) * Decimal(str(quantity))  # Convert rate and quantity to Decimal objects
                    # amount = float(rate)*float(quantity)
                    total += amount
                    LineItem(customer=invoice,
                            service=service,
                            # description=description,
                            quantity=quantity,
                            rate=rate,
                            amount=amount).save()
                    print(total)
            invoice.total_amount = total
            invoice.save()
            # messages.success(request, 'Invoice created successfully.')
            try:
                generate_PDF(request, id=invoice.id)
                # messages.success(request, 'Invoice created successfully.')
                # return JsonResponse({'success': True}) 
            except Exception as e:
                messages.error(request, f'An error occurred: {e}')
                return JsonResponse({'error': str(e)}, status=500)
                # print(f"********{e}********")
            return redirect('/invoice/')
    all_messages = messages.get_messages(request)
    context = {
        "title" : "Invoice Generator",
        "formset": formset,
        "form": form,
        'messages':all_messages
    }
    return render(request, 'invoice/invoice-create.html', context)




def editInvoice(request, id):
    """
    Invoice editing page.
    Only admin has the authority to read and make changes here.
    """

    invoice = get_object_or_404(Invoice, pk=id)
    try:
        if request.method == 'POST':
            # Update the attributes of the invoice object
            invoice.customer = request.POST.get('customer')
            invoice.customer_email = request.POST.get('customer_email')
            invoice.billing_address = request.POST.get('billing_address')
            # invoice.date = request.POST.get('date')
            invoice.date = datetime.strptime(request.POST.get('date'), '%Y-%m-%d')
            # invoice.due_date = request.POST.get('due_date')
            invoice.due_date = datetime.strptime(request.POST.get('due_date'), '%Y-%m-%d')
            invoice.message = request.POST.get('message')
            invoice.salesperson = request.POST.get('salesperson')
            invoice.manager = request.POST.get('manager')
            invoice.routing = request.POST.get('routing')
            invoice.total_amount = request.POST.get('total_amount')
            invoice.paid_amount = request.POST.get('paid_amount')
            invoice.balance = request.POST.get('balance')
            invoice.previous_balance = request.POST.get('previous_balance')
            # Update more fields as needed

            # Save the changes
            invoice.save()
            return redirect('/invoice/')
    except Exception as e:
            messages.error(request, f'An error occurred: {e}')
            # return JsonResponse({'error': str(e)}, status=500)

    # Get all messages and pass them to the template context
    all_messages = messages.get_messages(request)
    context = {
        'invoice': invoice,
        'invoice_id': id,
        'messages':all_messages
    }
    return render(request, 'invoice/invoice-edit.html', context)

# def editInvoice(request, id):
#     """
#     Invoice editing page.
#     Only admin has the authority to read and make changes here.
#     """

#     invoice = get_object_or_404(Invoice, id=id)

#     if request.method == 'POST':
#         form = InvoiceForm(request.POST, initial={'customer': invoice.customer,
#                                                    'customer_email': invoice.customer_email,
#                                                    'billing_address': invoice.billing_address,
#                                                    'date': invoice.date,
#                                                    'due_date': invoice.due_date,
#                                                    'message': invoice.message,
#                                                    'salesperson': invoice.salesperson,
#                                                    'manager': invoice.manager,
#                                                    'sale_types': invoice.sale_types,
#                                                    'routing': invoice.routing,
#                                                    'paid_amount': invoice.paid_amount,
#                                                    'balance':invoice.balance,
#                                                    'previous_balance':invoice.previous_balance

#                                                    })
#         if form.is_valid():
#             # Update the existing invoice with the form data
#             invoice.customer = form.cleaned_data['customer']
#             invoice.customer_email = form.cleaned_data['customer_email']
#             invoice.billing_address = form.cleaned_data['billing_address']
#             invoice.date = form.cleaned_data['date']
#             invoice.due_date = form.cleaned_data['due_date']
#             invoice.message = form.cleaned_data['message']
#             invoice.salesperson = form.cleaned_data['salesperson']
#             invoice.manager = form.cleaned_data['manager']
#             invoice.sale_types = form.cleaned_data['sale_types']
#             invoice.routing = form.cleaned_data['routing']
#             invoice.paid_amount = form.cleaned_data['paid_amount']
#             invoice.balance = form.cleaned_data['balance']
#             invoice.previous_balance = form.cleaned_data['previous_balance']

#             invoice.save()
#             messages.success(request, 'Invoice updated successfully.')
#             return redirect('/invoice/')
#     else:
#         form = InvoiceForm(initial={'customer': invoice.customer,
#                                      'customer_email': invoice.customer_email,
#                                      'billing_address': invoice.billing_address,
#                                      'date': invoice.date,
#                                      'due_date': invoice.due_date,
#                                      'message': invoice.message,
#                                      'salesperson': invoice.salesperson,
#                                      'manager': invoice.manager,
#                                      'sale_types': invoice.sale_types,
#                                      'routing': invoice.routing,
#                                      'paid_amount': invoice.paid_amount,
#                                      'balance':invoice.balance,
#                                      'previous_balance':invoice.previous_balance
#                                      })

#     context = {
#         "title" : "Edit Invoice",
#         "form": form,
#         "invoice_id":id,
#     }
#     return render(request, 'invoice/invoice-edit.html', context)

def delete_invoice(request, id):
    invoice = get_object_or_404(Invoice, id=id)
    try:
        if request.method == 'POST':
            invoice.delete()
            return redirect('/invoice/')
    except Exception as e:
                messages.error(request, f'An error occurred: {e}')
                return JsonResponse({'error': str(e)}, status=500)
    
    return render(request, 'invoice/delete_invoice_confirm.html', {'invoice': invoice})


def view_PDF(request, id=None):
    invoice = get_object_or_404(Invoice, id=id)
    lineitem = invoice.lineitem_set.all()


    # Extract day, month, and year components from the due_date
    invoice_date = invoice.date
    day = invoice_date.day
    month = invoice_date.month
    year = invoice_date.year


    context = {
        "company": {
            "name": "Sh.A.Rehman Traders",
            "address" :"Main Bazaar Piplan (Mianwali)",
            "phone": "03006084456",
            # "email": "asad@gmail.com",
        },
        "invoice_id": f"{invoice.id}{day}{month}{year}",
        "invoice_total": invoice.total_amount,
        "customer": invoice.customer,
        "customer_email": invoice.customer_email,
        "date": invoice_date,
        "due_date": invoice.due_date,
        "billing_address": invoice.billing_address,
        "message": invoice.message,
        "Sale_Type":invoice.sale_types,
        "lineitem": lineitem,

    }
    return render(request, 'invoice/pdf_template.html', context)

def generate_PDF(request, id):
    # Use False instead of output path to save pdf to a variable
    config = pdfkit.configuration(wkhtmltopdf='C:\\Program Files\\wkhtmltopdf\\bin\wkhtmltopdf.exe')
    pdf = pdfkit.from_url(request.build_absolute_uri(reverse('invoice:invoice-detail', args=[id])), False, configuration=config)
    response = HttpResponse(pdf,content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="invoice.pdf"'

    return response

# def change_status(request):
#     return redirect('invoice:invoice-list')

def view_404(request,  *args, **kwargs):

    return redirect('invoice:invoice-list')

from django.http import JsonResponse


# def get_details_customer(request, shop_name):
#     shop = Shop.objects.get(id=shop_name)
#     data = {
#         'name': shop.name,
#         'address': shop.address,
#         'owner_number': shop.owner_number,
#         'owner_cnic': shop.owner_cnic,
#     }
#     return JsonResponse(data)


def get_details_customer(request, shop_name):
    try:
        shop = Shop.objects.get(name=shop_name)
        data = {
            'name': shop.name,
            'address': shop.address,
            'owner_number': shop.owner_number,
            'owner_cnic': shop.owner_cnic,
        }
        return JsonResponse(data)
    except Shop.DoesNotExist:
        return JsonResponse({'error': 'Shop not found'}, status=404)

# Showing httpresponse when i hit /invoice/get_details_customer/zulfiqarali in browser  
# def get_details_customer(request, shop_name):
#     try:
#         shop = get_object_or_404(Shop, name=shop_name)
#         data = {
#             'name': shop.name,
#             'address': shop.address,
#             'owner_number': shop.owner_number,
#             'owner_cnic': shop.owner_cnic,
#         }
#         return HttpResponse(json.dumps(data), content_type='application/json')
#     except Http404:
#         return HttpResponse(json.dumps({'error': 'Shop not found'}), status=404, content_type='application/json')

    
def get_product_rate(request, product_name):
    try:
        product = Product.objects.get(product_name=product_name)
        data = {
            'rate': product.price,  # Assuming 'price' is the rate field
        }
        return JsonResponse(data)
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)
    
from django.db import transaction
from django.db.models import F

def create_recovery(request, id):
     thank=False
     invoice = get_object_or_404(Invoice, id=id)
     if request.method=="POST":
        customer = invoice.customer
        # total_amount = invoice.total_amount
        total_amount = Decimal(invoice.total_amount)
        total_amount= Decimal(total_amount)
        date = invoice.date.strftime('%Y-%m-%d')
        balance = invoice.balance
        existing_paid_amount = Decimal(invoice.paid_amount)
        # existing_paid_amount=Invoice.paid_amount
        new_paid_amount_str = request.POST.get('new_paid_amount', '')
        new_paid_amount = Decimal(new_paid_amount_str)
        current_balance = (invoice.total_amount - (new_paid_amount+existing_paid_amount))
        recovery = Recovery(customer_name=customer, total_amount=total_amount, date=date, balance=balance,new_paid_amount=new_paid_amount,current_balance=current_balance)
        recovery.save()
        # Update the invoice's paid_amount field by adding new_paid_amount to it
        with transaction.atomic():
            Invoice.objects.filter(id=id).update(
                paid_amount=F('paid_amount') + new_paid_amount,
                # previous_balance=F('previous_balance')+ balance,
                previous_balance=balance,
                balance=current_balance,
            )

        thank=True
        return redirect('invoice:invoice-list')
     
     return render(request, 'invoice/recovery_form.html', {'thank':thank,'invoice_id': id})

def recovery_list(request):
    recoveryies_fetched = Recovery.objects.all().order_by("-id")
    # print(days_fetched)
    return render(request, 'invoice/Recovery_list.html', {'recoveryies_fetched': recoveryies_fetched})

#=== edit_route views function
def edit_recovery(request, pk):
    recovery = get_object_or_404(Recovery, pk=pk)
    if request.method == 'POST':
        form = RecoveryForm(request.POST, instance=recovery)
        if form.is_valid():
            form.save()
            return redirect('invoice:recovery-list')
    else:
        form = RecoveryForm(instance=recovery)
    return render(request, 'invoice/edit_recovery.html', {'form': form})

#=== delete_route views function
def delete_recovery(request, pk):
    recovery = get_object_or_404(Recovery, pk=pk)
    if request.method == 'POST':
        recovery.delete()
        return redirect('invoice:recovery-list')
    return render(request, 'invoice/delete_recovery.html', {'route': recovery})

from openpyxl import Workbook

# def generate_invoice_excel_report(request):
#     if request.method == 'GET':
#             # Handle GET request (render a form, for example)
#             return render(request, 'invoice/main_invoice_excel.html', context={})
#     elif request.method == 'POST':
#         # Fetch all invoices from the database
#         invoices = Invoice.objects.all()
#         # Create a new Excel workbook
#         wb = Workbook()
#         # Add a worksheet for the report
#         ws = wb.active
#         ws.title = "Invoice Report"
#         # Write headers
#         headers = [
#             "Customer", "Customer Email", "Billing Address", "Date", "Due Date", "Message",
#             "Total Amount", "Salesperson", "Manager", "Routing", "Paid Amount", "Balance",
#             "Previous Balance", "Sale Types", "Status"
#         ]
#         ws.append(headers)
#         # Write invoice data rows
#         for invoice in invoices:
#             row = [
#                 invoice.customer, invoice.customer_email, invoice.billing_address,
#                 invoice.date, invoice.due_date, invoice.message,
#                 invoice.total_amount, invoice.salesperson, invoice.manager,
#                 invoice.routing, invoice.paid_amount, invoice.balance,
#                 invoice.previous_balance, invoice.sale_types, invoice.status
#             ]
#             ws.append(row)
#         # Create HTTP response with Excel content type
#         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename=invoice_report.xlsx'
#         # Save the workbook to the HTTP response
#         wb.save(response)

#         return response
    
def generate_invoice_excel_report(request):
    if request.method == 'GET':
        # Handle GET request (render a form, for example)
        return render(request, 'invoice/main_invoice_excel.html', context={})
    elif request.method == 'POST':
        report_type = request.POST.get('report_type')
        
        if report_type == 'all':
            return generate_report_all_data(request)
        elif report_type == 'by_manager':
            manager_name = request.POST.get('manager_name')
            return generate_report_by_manager(request, manager_name)
        elif report_type == 'by_salesman':
            salesman_name = request.POST.get('salesman_name')
            return generate_report_by_salesman(request, salesman_name)
        elif report_type == 'between_dates':
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            return generate_report_between_dates(request, start_date, end_date)
        else:
            # Default behavior if no report type specified
            return HttpResponse("Invalid report type")
        
def generate_report_all_data(request):
    # Fetch all invoices from the database
    invoices = Invoice.objects.all()
    return generate_excel_report(invoices)


def generate_report_by_manager(request, manager_name):
    # Fetch invoices filtered by manager from the database
    invoices = Invoice.objects.filter(manager=manager_name)
    return generate_excel_report(invoices)

def generate_report_by_salesman(request, salesman_name):
    # Fetch invoices filtered by salesman from the database
    invoices = Invoice.objects.filter(salesperson=salesman_name)
    return generate_excel_report(invoices)

def generate_report_between_dates(request, start_date, end_date):
    # Fetch invoices between the specified dates from the database
    invoices = Invoice.objects.filter(date__range=[start_date, end_date])
    return generate_excel_report(invoices)

def generate_excel_report(invoices):
    # Create a new Excel workbook
    wb = Workbook()
    # Add a worksheet for the report
    ws = wb.active
    ws.title = "Invoice Report"
    # Write headers
    headers = [
        "Customer", "Customer Email", "Billing Address", "Date", "Due Date", "Message",
        "Total Amount", "Salesperson", "Manager", "Routing", "Paid Amount", "Balance",
        "Previous Balance", "Sale Types", "Status"
    ]
    ws.append(headers)
    # Write invoice data rows
    for invoice in invoices:
        row = [
            invoice.customer, invoice.customer_email, invoice.billing_address,
            invoice.date, invoice.due_date, invoice.message,
            invoice.total_amount, invoice.salesperson, invoice.manager,
            invoice.routing, invoice.paid_amount, invoice.balance,
            invoice.previous_balance, invoice.sale_types, invoice.status
        ]
        ws.append(row)
    # Create HTTP response with Excel content type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=invoice_report.xlsx'
    # Save the workbook to the HTTP response
    wb.save(response)

    return response








    
# name,address,owner_number,owner_cnic
def generate_shop_excel_report(request):
    if request.method == 'GET':
            # Handle GET request (render a form, for example)
            return render(request, 'invoice/main_shop_excel.html', context={})
    elif request.method == 'POST':
        # Fetch all invoices from the database
        shops = Shop.objects.all()
        # Create a new Excel workbook
        wb = Workbook()
        # Add a worksheet for the report
        ws = wb.active
        ws.title = "Shop List Report "
        # Write headers
        headers = [
            "name", "address", "owner_number", "owner_cnic"
        ]
        ws.append(headers)
        # Write invoice data rows
        for shop in shops:
            row = [
                shop.name, shop.address, shop.owner_number,
                shop.owner_cnic
            ]
            ws.append(row)
        # Create HTTP response with Excel content type
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=shops_report.xlsx'
        # Save the workbook to the HTTP response
        wb.save(response)

        return response
    # name,address,owner_number,owner_cnic
def generate_recoveries_excel_report(request):
    if request.method == 'GET':
            # Handle GET request (render a form, for example)
            return render(request, 'invoice/main_recoveries_excel.html', context={})
    elif request.method == 'POST':
        # Fetch all invoices from the database
        recoveries = Recovery.objects.all()
        # Create a new Excel workbook
        wb = Workbook()
        # Add a worksheet for the report
        ws = wb.active
        ws.title = "Shop List Report "
        # Write headers
        headers = [
            "Customer/Shop Name", "Total Sale Amount", "Order Date", "Previous Balance",'Recovery/Updated Date','Recovered Amount','Current Balance(payable)'
        ]
        ws.append(headers)
        # Write invoice data rows
        for recovery in recoveries:
            row = [
                recovery.customer_name, recovery.total_amount, recovery.date,
                recovery.balance,recovery.updated_date,recovery.new_paid_amount,recovery.current_balance
            ]
            ws.append(row)
        # Create HTTP response with Excel content type
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Recoveries_report.xlsx'
        # Save the workbook to the HTTP response
        wb.save(response)

        return response